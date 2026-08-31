"""Tabular and structured data: CSV, TSV, XLSX, JSON, YAML, XML, TOML."""

import io
import csv
import json
import html as htmlmod
from pathlib import Path

from engine import rule, ConvertError

csv.field_size_limit(10 ** 8)


# --------------------------------------------------------------------------
# helpers: everything goes through "rows" (list of lists) or "obj" (json-ish)
# --------------------------------------------------------------------------

def _sniff(path):
    sample = Path(path).read_text(encoding="utf-8-sig", errors="replace")[:8192]
    if Path(path).suffix.lower() == ".tsv":
        return "\t"
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        return "\t" if sample.count("\t") > sample.count(",") else ","


def _read_rows(path):
    delim = _sniff(path)
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return [r for r in csv.reader(f, delimiter=delim)]


def _write_rows(path, rows, delim=","):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f, delimiter=delim).writerows(rows)


def _rows_to_objs(rows):
    if not rows:
        return []
    head = [h.strip() or "col%d" % (i + 1) for i, h in enumerate(rows[0])]
    return [dict(zip(head, r + [""] * (len(head) - len(r)))) for r in rows[1:]]


def _objs_to_rows(obj):
    """Best-effort table out of any JSON-ish structure."""
    if isinstance(obj, dict):
        for v in obj.values():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                obj = v
                break
        else:
            obj = [obj]
    if not isinstance(obj, list):
        obj = [obj]
    if obj and all(isinstance(x, dict) for x in obj):
        keys = []
        for d in obj:
            for k in d:
                if k not in keys:
                    keys.append(k)
        rows = [keys]
        for d in obj:
            rows.append([_flat(d.get(k, "")) for k in keys])
        return rows
    if obj and all(isinstance(x, (list, tuple)) for x in obj):
        return [list(x) for x in obj]
    return [["value"]] + [[_flat(x)] for x in obj]


def _flat(v):
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return "" if v is None else v


def _rows_to_html(rows, title):
    def cells(r, tag):
        return "".join("<%s>%s</%s>" % (tag, htmlmod.escape(str(c)), tag) for c in r)
    head = "<thead><tr>%s</tr></thead>" % cells(rows[0], "th") if rows else ""
    body = "".join("<tr>%s</tr>" % cells(r, "td") for r in rows[1:])
    return ("<!doctype html><meta charset='utf-8'><title>%s</title>"
            "<style>body{font:14px/1.5 Segoe UI,sans-serif;margin:2em}"
            "table{border-collapse:collapse}th,td{border:1px solid #ccc;padding:.35em .7em}"
            "th{background:#f2f2f4;text-align:left}</style>"
            "<table>%s<tbody>%s</tbody></table>"
            % (htmlmod.escape(title), head, body))


def _rows_to_md(rows):
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    rows = [list(r) + [""] * (width - len(r)) for r in rows]
    esc = lambda c: str(c).replace("|", "\\|").replace("\n", " ")
    out = ["| " + " | ".join(esc(c) for c in rows[0]) + " |",
           "| " + " | ".join("---" for _ in rows[0]) + " |"]
    out += ["| " + " | ".join(esc(c) for c in r) + " |" for r in rows[1:]]
    return "\n".join(out) + "\n"


def _load_obj(path):
    ext = Path(path).suffix.lower().lstrip(".")
    text = Path(path).read_text(encoding="utf-8-sig", errors="replace")
    if ext == "json":
        return json.loads(text)
    if ext in ("yaml", "yml"):
        import yaml
        return yaml.safe_load(text)
    if ext == "xml":
        import xmltodict
        return xmltodict.parse(text)
    if ext == "toml":
        import tomli
        return tomli.loads(text)
    if ext in ("jsonl", "ndjson"):
        return [json.loads(l) for l in text.splitlines() if l.strip()]
    raise ConvertError("cannot read %s as structured data" % ext)


# --------------------------------------------------------------------------
# spreadsheets
# --------------------------------------------------------------------------

try:
    import openpyxl

    @rule("xlsx xlsm xltx", "csv tsv json html md txt", cost=6, label="spreadsheet out")
    def xlsx_out(src, out, dst):
        wb = openpyxl.load_workbook(str(src), data_only=True, read_only=True)
        rows = []
        for i, ws in enumerate(wb.worksheets):
            if i and rows:
                rows.append([])
                rows.append(["# sheet: " + ws.title])
            for r in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else c for c in r])
        wb.close()
        _emit_rows(rows, out, dst, Path(src).stem)

    @rule("csv tsv json", "xlsx", cost=6, label="to spreadsheet")
    def to_xlsx(src, out, dst):
        ext = Path(src).suffix.lower().lstrip(".")
        rows = _objs_to_rows(_load_obj(src)) if ext == "json" else _read_rows(src)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = Path(src).stem[:31] or "Sheet1"
        for r in rows:
            ws.append([_flat(c) for c in r])
        if rows:
            for c in ws[1]:
                c.font = openpyxl.styles.Font(bold=True)
            ws.freeze_panes = "A2"
            for col in ws.columns:
                width = max((len(str(c.value or "")) for c in col[:200]), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 8), 60)
        wb.save(str(out))

except ImportError:
    pass


def _emit_rows(rows, out, dst, title):
    out = Path(out)
    if dst == "csv":
        _write_rows(out, rows)
    elif dst == "tsv":
        _write_rows(out, rows, "\t")
    elif dst == "json":
        out.write_text(json.dumps(_rows_to_objs(rows), indent=2, ensure_ascii=False,
                                  default=str), encoding="utf-8")
    elif dst == "html":
        out.write_text(_rows_to_html(rows, title), encoding="utf-8")
    elif dst == "md":
        out.write_text(_rows_to_md(rows), encoding="utf-8")
    elif dst == "txt":
        out.write_text("\n".join("\t".join(str(c) for c in r) for r in rows),
                       encoding="utf-8")
    elif dst in ("yaml", "yml"):
        import yaml
        out.write_text(yaml.safe_dump(_rows_to_objs(rows), allow_unicode=True,
                                      sort_keys=False), encoding="utf-8")
    else:
        raise ConvertError("cannot write %s here" % dst)


# --------------------------------------------------------------------------
# delimited text
# --------------------------------------------------------------------------

@rule("csv tsv", "csv tsv json html md txt yaml yml", cost=6, label="csv out")
def csv_out(src, out, dst):
    _emit_rows(_read_rows(src), out, dst, Path(src).stem)


# --------------------------------------------------------------------------
# structured data
# --------------------------------------------------------------------------

STRUCT = "json yaml yml xml toml jsonl ndjson"


@rule(STRUCT, "json yaml yml xml txt", cost=6, label="structured data")
def struct_out(src, out, dst):
    obj = _load_obj(src)
    out = Path(out)
    if dst == "json":
        out.write_text(json.dumps(obj, indent=2, ensure_ascii=False, default=str),
                       encoding="utf-8")
    elif dst in ("yaml", "yml"):
        import yaml
        out.write_text(yaml.safe_dump(obj, allow_unicode=True, sort_keys=False,
                                      default_flow_style=False), encoding="utf-8")
    elif dst == "xml":
        import xmltodict
        if isinstance(obj, list):
            obj = {"root": {"item": obj}}       # a list needs a repeating element
        elif not isinstance(obj, dict) or len(obj) != 1:
            obj = {"root": obj}
        out.write_text(xmltodict.unparse(obj, pretty=True), encoding="utf-8")
    elif dst == "txt":
        out.write_text(json.dumps(obj, indent=2, ensure_ascii=False, default=str),
                       encoding="utf-8")


@rule("json yaml yml xml jsonl ndjson", "csv tsv html md", cost=7, label="data to table")
def struct_to_table(src, out, dst):
    _emit_rows(_objs_to_rows(_load_obj(src)), out, dst, Path(src).stem)


try:
    import tomli_w

    @rule("json yaml yml toml", "toml", cost=7, label="to toml")
    def to_toml(src, out, dst):
        obj = _load_obj(src)
        if not isinstance(obj, dict):
            obj = {"root": obj}
        Path(out).write_bytes(tomli_w.dumps(obj).encode("utf-8"))

except ImportError:
    pass
